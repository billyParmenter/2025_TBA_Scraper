import requests
import pandas as pd
import os
import configparser
import openpyxl
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import ttk, messagebox
import threading
import sys

#
# --- CONFIGURATION & UTILITY FUNCTIONS ---
#

# Determine base path: use _MEIPASS if running as a bundled executable.
if hasattr(sys, '_MEIPASS'):
    base_path = sys._MEIPASS
else:
    base_path = os.path.abspath(".")

# Load config.ini
config_path = os.path.join(base_path, "config.ini")
config = configparser.ConfigParser()
found = config.read(config_path)
if not found or 'tba' not in config:
    raise Exception(f"Config file not found or missing [tba] section at {config_path}")
API_KEY = config.get("tba", "api_key")

# Global settings
BASE_URL = "https://www.thebluealliance.com/api/v3"
YEAR = 2025
OUTPUT_FILENAME = "matches.xlsx"

headers = {
    'accept': 'application/json',
    'X-TBA-Auth-Key': API_KEY,
}

# Keys to ignore in the match-level common data.
IGNORE_TOP_KEYS = {
    "actual_time",
    "comp_level",
    "match_number",
    "key",
    "post_result_time",
    "predicted_time",
    "set_number",
    "time",
    "videos"
}

# Reef sub-keys to ignore in autoReef and teleopReef
IGNORE_REEF_KEYS = {"botRow", "midRow", "topRow"}


def fetch_available_events(year=YEAR):
    """Fetch a list of events for the given year from TBA."""
    url = f"{BASE_URL}/events/{year}"
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        return response.json()  # List of event dicts.
    else:
        messagebox.showerror("Error", f"Error fetching events: {response.status_code}")
        return []


def process_score_breakdown(sb):
    """
    Flatten the score_breakdown for an alliance.
    For keys 'autoReef' and 'teleopReef', flatten their nested dictionary
    excluding IGNORE_REEF_KEYS.
    Also skip the 'totalPoints' key.
    (No prefix is added.)
    """
    flat = {}
    for key, value in sb.items():
        if key == "totalPoints":
            continue  # skip totalPoints
        if key in ['autoReef', 'teleopReef'] and isinstance(value, dict):
            for sub_key, sub_val in value.items():
                if sub_key not in IGNORE_REEF_KEYS:
                    flat[f"{key}_{sub_key}"] = sub_val
        else:
            flat[key] = value
    return flat


def process_match(match):
    """
    Process a single match record and return a list of two dictionaries
    (one per alliance) containing:
      - Common match-level fields (excluding IGNORE_TOP_KEYS)
      - A custom "match" identifier is built from comp_level and match_number.
      - Flattened score_breakdown fields for that alliance, skipping reef keys and totalPoints.
    """
    # Skip if alliances or score_breakdown is None (incomplete match).
    if match.get("alliances") is None or match.get("score_breakdown") is None:
        return []

    # Build common data.
    common_data = {}
    for k, v in match.items():
        if k not in IGNORE_TOP_KEYS and k not in ['alliances', 'score_breakdown']:
            common_data[k] = v

    # Create a custom "match" identifier.
    comp_level = match.get("comp_level", "")
    match_number = match.get("match_number", "")
    common_data["match"] = f"{comp_level} {match_number}"

    alliances = match["alliances"]
    score_breakdown = match["score_breakdown"]

    output_rows = []
    for alliance_color in ['red', 'blue']:
        row = common_data.copy()
        row["alliance"] = alliance_color

        # Process alliance-specific data.
        alliance_data = alliances.get(alliance_color, {})
        team_keys = alliance_data.get("team_keys", [])
        # Remove "frc" prefix and cast to int.
        try:
            row["bot1"] = int(team_keys[0].replace("frc", "")) if len(team_keys) > 0 else None
        except:
            row["bot1"] = None
        try:
            row["bot2"] = int(team_keys[1].replace("frc", "")) if len(team_keys) > 1 else None
        except:
            row["bot2"] = None
        try:
            row["bot3"] = int(team_keys[2].replace("frc", "")) if len(team_keys) > 2 else None
        except:
            row["bot3"] = None

        # Include the alliance's score if available.
        row["alliance_score"] = alliance_data.get("score", None)

        # Flatten the alliance-specific score breakdown.
        sb = score_breakdown.get(alliance_color, {})
        if sb:
            flat_sb = process_score_breakdown(sb)
            row.update(flat_sb)

        output_rows.append(row)

    return output_rows


def fetch_event_matches(event_key, error_list):
    """Fetch match data for a specific event, process it, and collect errors in error_list."""
    url = f"{BASE_URL}/event/{event_key}/matches"
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        try:
            matches = response.json()
        except Exception as e:
            error_msg = (
                f"Error parsing JSON for event {event_key}:\n{e}\n"
                f"Response text: {response.text}"
            )
            error_list.append(error_msg)
            return []
        rows = []
        for i, match in enumerate(matches):
            try:
                processed = process_match(match)
                rows.extend(processed)
            except Exception as e:
                msg = (
                    f"Error processing match index {i} for event {event_key}:\n"
                    f"Match data: {match}\n\n"
                    f"Exception: {e}\n\n"
                    f"URL: {url}\n\n"
                    "This match was skipped."
                )
                error_list.append(msg)
                continue  # Skip the problematic match
        return rows
    else:
        error_msg = (
            f"Error fetching matches for {event_key}:\n"
            f"Status code: {response.status_code}\n"
            f"Response: {response.text}"
        )
        error_list.append(error_msg)
        return []


def reorder_columns(df):
    """
    Reorder columns so that "event_key" is first,
    followed by "match" and then "winning_alliance" (if present),
    with the remaining columns following.
    """
    cols = list(df.columns)
    desired_order = []
    for col in ["event_key", "match", "winning_alliance"]:
        if col in cols:
            desired_order.append(col)
            cols.remove(col)
    new_order = desired_order + cols
    return df[new_order]


def auto_adjust_excel_columns(filename):
    """Auto-adjust column widths in an Excel file using openpyxl."""
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    wb.save(filename)


def load_existing_data():
    """Load the existing Excel file if it exists, or return an empty DataFrame."""
    if os.path.exists(OUTPUT_FILENAME):
        return pd.read_excel(OUTPUT_FILENAME)
    else:
        return pd.DataFrame()


def update_file(new_data_df, mode="replace"):
    """
    Update the Excel file with new_data_df.
    If mode == "replace", overwrite the file with new_data_df.
    If mode == "update", remove rows for those event_keys in new_data_df and then append new_data_df.
    """
    if mode == "replace" or not os.path.exists(OUTPUT_FILENAME):
        df = new_data_df
    else:
        existing_df = pd.read_excel(OUTPUT_FILENAME)
        keys_to_replace = new_data_df["event_key"].unique()
        updated_df = existing_df[~existing_df["event_key"].isin(keys_to_replace)]
        df = pd.concat([updated_df, new_data_df], ignore_index=True)

    df = reorder_columns(df)
    df.to_excel(OUTPUT_FILENAME, index=False)
    auto_adjust_excel_columns(OUTPUT_FILENAME)


#
# --- UI FUNCTIONS ---
#

def update_file_events_list():
    """Update the UI list of events already in the file."""
    df = load_existing_data()
    file_events_listbox.delete(0, tk.END)
    if not df.empty and "event_key" in df.columns:
        event_keys = sorted(df["event_key"].unique().tolist())
        for key in event_keys:
            file_events_listbox.insert(tk.END, key)


def load_all_events():
    """
    Fetch match data for all available events from TBA, replace the file,
    and collect any errors in a list for a single popup at the end.
    """
    error_list = []
    all_rows = []

    for event in available_events:
        event_key = event["key"]
        print(f"Fetching matches for {event_key}...")
        rows = fetch_event_matches(event_key, error_list)  # pass error_list
        all_rows.extend(rows)

    if all_rows:
        new_df = pd.DataFrame(all_rows)
        update_file(new_df, mode="replace")
        messagebox.showinfo("Success", "File replaced with all TBA events data.")
        update_file_events_list()
    else:
        messagebox.showerror("Error", "No data fetched.")

    # Show all errors at the end (if any)
    if error_list:
        combined_errors = "\n\n".join(error_list)
        messagebox.showwarning("Some Events Had Errors", combined_errors)


def add_replace_selected():
    """
    Fetch match data for selected events, update the file,
    and collect any errors in a list for a single popup at the end.
    """
    selected_indices = available_events_listbox.curselection()
    if not selected_indices:
        messagebox.showwarning("No Selection", "No events selected.")
        return

    error_list = []
    all_rows = []

    for idx in selected_indices:
        event = available_events[idx]
        event_key = event["key"]
        print(f"Fetching matches for {event_key}...")
        rows = fetch_event_matches(event_key, error_list)  # pass error_list
        all_rows.extend(rows)

    if all_rows:
        new_df = pd.DataFrame(all_rows)
        update_file(new_df, mode="update")
        messagebox.showinfo("Success", "Selected events updated in file.")
        update_file_events_list()
    else:
        messagebox.showerror("Error", "No data fetched for selected events.")

    # Show all errors at the end (if any)
    if error_list:
        combined_errors = "\n\n".join(error_list)
        messagebox.showwarning("Some Events Had Errors", combined_errors)


def refresh_available_events():
    """Fetch available events from TBA and populate the listbox."""
    global available_events
    all_events = fetch_available_events(YEAR)

    # Optionally, filter out 'tempclone' or test events if you want:
    # all_events = [e for e in all_events if 'tempclone' not in e['key']]

    available_events = all_events
    available_events_listbox.delete(0, tk.END)
    for event in available_events:
        display_text = f"{event['key']} - {event.get('name', 'No Name')}"
        available_events_listbox.insert(tk.END, display_text)


#
# --- LOADING INDICATOR & THREAD WRAPPERS ---
#

def start_loading():
    """Show and start the progress bar."""
    progress_bar.grid()  # Make it visible.
    progress_bar.start(10)  # Start animation with 10ms interval.

def stop_loading():
    """Stop and hide the progress bar."""
    progress_bar.stop()
    progress_bar.grid_remove()

def threaded_load_all_events():
    """Run load_all_events in a background thread with a loading indicator."""
    def task():
        load_all_events()
        root.after(0, stop_loading)
    root.after(0, start_loading)
    threading.Thread(target=task, daemon=True).start()

def threaded_add_replace_selected():
    """Run add_replace_selected in a background thread with a loading indicator."""
    def task():
        add_replace_selected()
        root.after(0, stop_loading)
    root.after(0, start_loading)
    threading.Thread(target=task, daemon=True).start()


#
# --- SET UP THE TKINTER UI ---
#

root = tk.Tk()
root.title("TBA Match Data Updater")

# Frames for layout.
frame_top = ttk.Frame(root, padding="10")
frame_top.grid(row=0, column=0, sticky="nsew")
frame_bottom = ttk.Frame(root, padding="10")
frame_bottom.grid(row=1, column=0, sticky="nsew")

# Available Events from TBA.
ttk.Label(frame_top, text="Available Events from TBA:").grid(row=0, column=0, sticky="w")
available_events_listbox = tk.Listbox(frame_top, selectmode=tk.MULTIPLE, width=50, height=10)
available_events_listbox.grid(row=1, column=0, padx=5, pady=5)

# Buttons for actions.
btn_replace_all = ttk.Button(frame_top, text="Load All Events (Replace File)", command=threaded_load_all_events)
btn_replace_all.grid(row=2, column=0, pady=(5,2), sticky="ew")

btn_add_replace = ttk.Button(frame_top, text="Add/Replace Selected Events", command=threaded_add_replace_selected)
btn_add_replace.grid(row=3, column=0, pady=(2,5), sticky="ew")

# Create an indeterminate progress bar (initially hidden).
progress_bar = ttk.Progressbar(frame_top, mode="indeterminate")
progress_bar.grid(row=4, column=0, pady=(5,5), sticky="ew")
progress_bar.grid_remove()  # Hide it initially.

# Events already in the file.
ttk.Label(frame_bottom, text="Events already in the file:").grid(row=0, column=0, sticky="w")
file_events_listbox = tk.Listbox(frame_bottom, width=50, height=5)
file_events_listbox.grid(row=1, column=0, padx=5, pady=5)

# Refresh button for available events.
btn_refresh = ttk.Button(frame_bottom, text="Refresh Available Events", command=refresh_available_events)
btn_refresh.grid(row=2, column=0, pady=(5,0), sticky="ew")

# Set grid weights.
root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)
root.rowconfigure(1, weight=0)

# Global variable to store available events.
available_events = []

# Initially load available events and update file events list.
refresh_available_events()
update_file_events_list()

root.mainloop()
